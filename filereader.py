import openpyxl as op
from openpyxl.utils import get_column_letter

class FileReader():
    def __init__(self,fname):
        self.cwd = ''
        self.fname = fname
        self.excelName = fname + '.xlsx'
        self.filename = fname + '.txt'
        self.filename_answer = fname + ' answers.txt'
        self.mc_list = []
        self.tf_list = []
        self.fill_list = []
        self.answer_list = []
        self.question_fill()
        self.answer_fill()
        self.excelWriter()

    def openFile(self): # opens the raw file and lets the lines be accessed
        with open(self.filename) as openFile:
            self.raw_lines = openFile.read().splitlines()

    def break_points(self): # sets the ranges of lines for each type of question
        self.mc = self.raw_lines.index('Multiple Choice:')
        self.tf = self.raw_lines.index('True/False:')
        try:
            self.fill = self.raw_lines.index('Fill in the blank:')
        except:
            self.fill = self.raw_lines.index('Fill in the Blank:')

    def mcParser(self):
        temp_list = []
        for line in range(self.mc + 1,self.tf):
            b_space = self.raw_lines[line].index(' ')
            n_space = self.raw_lines[line].index('.')
            line_value = self.raw_lines[line][:n_space]
            line_text = self.raw_lines[line][b_space + 1:]
            try:
                if int(line_value) in range(0,99):
                    temp_list.append('*' + line_text)
                    for item in temp_list:
                        self.mc_list.append(item)
                    temp_list = []
            except:
                temp_list.append(line_text)
        for item in temp_list:
            self.mc_list.append(item)

    def tfParser(self):
        for line in range(self.tf + 1,self.fill):
            w_space = self.raw_lines[line].index(' ',6)
            text = self.raw_lines[line][w_space + 1:]
            self.tf_list.append(text)

    def fillParser(self):
        for line in range(self.fill + 1,len(self.raw_lines)):
            w_space = self.raw_lines[line].index(' ')
            text = self.raw_lines[line][w_space + 1:]
            self.fill_list.append(text)

    def question_fill(self):
        self.openFile()
        self.break_points()
        self.mcParser()
        self.tfParser()
        self.fillParser()

    def answer_fill(self):
        with open(self.filename_answer) as openFile:
            raw_answers = openFile.read().splitlines()
            for item in raw_answers:
                w_space = item.index(' ')
                answer = item[w_space + 1:].title()
                self.answer_list.append(answer)

    def excelWriter(self):
        wb = op.Workbook()
        ws = wb.active
        ws.append(['Q','A','B','C','D','Answer'])
        x = 1
        y = 1
        for item in self.mc_list:
            if item[0] == '*':
                x = 1
                y += 1
                x_let = get_column_letter(x)
                ws[x_let + str(y)] = item[1:]
            else:
                x += 1
                x_let = get_column_letter(x)
                ws[x_let + str(y)] = item
        for item in self.tf_list:
            x = 1
            y += 1
            x_let = get_column_letter(x)
            x_let_true = get_column_letter(x + 1)
            x_let_false = get_column_letter(x + 2)
            ws[x_let + str(y)] = item
            ws[x_let_true + str(y)] = 'True'
            ws[x_let_false + str(y)] = 'False'
        for item in self.fill_list:
            x = 1
            y += 1
            x_let = get_column_letter(x)
            ws[x_let + str(y)] = item
        y = 1
        x = 6
        for item in self.answer_list:
            y += 1
            x_let = get_column_letter(x)
            ws[x_let + str(y)] = item
        wb.save(self.excelName)
        wb.close()
        wb = None
        ws = None

    def moveFile(self): #Used to move the file already processed by wNames
        import os
        file_list = [self.excelName,self.filename,self.filename_answer]
        inipath = os.getcwd() + chr(92)
        path = os.getcwd() + chr(92) + self.fname + chr(92)
        access_rights = 0o755
        try:
            os.mkdir(path, access_rights)
        except:
            pass
        for item in file_list:
            os.rename(inipath + self.fname,item)